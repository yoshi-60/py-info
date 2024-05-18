#!/usr/bin/env python3

import sys
import os
import pathlib
import yaml
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries, coordinate_from_string, column_index_from_string
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.properties import Outline
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

def info_xlsx(fxlsx):
  # Excel book open
  if os.path.isfile(fxlsx):
    wb = load_workbook(filename=fxlsx, read_only=True)
  else:
    print (f'File {fxlsx} Not found.')
    return

  # Excel Info
  print(f'Excel_file {fxlsx}.')
  print(f'  Sheets')
  i = 1
  for sheet in wb:
    print(f'    {i:3d} sheet_name: {(sheet.title+",").ljust(25)} last_row: {sheet.max_row:6d}, last_column: {sheet.max_column:3d}',end='')
    print(f' ( {get_column_letter(sheet.max_column)+str(sheet.max_row)} )')
    i = i+1
  return

def edit_xlsx(fxlsx, fyml):
  # open yaml file
  with open ( fyml, 'r') as yml:
    xlparam = yaml.safe_load(yml)

  # get parameters
  #print( xlparam['sheets'] )|
  #sheets_num = len (x]param[' sheets' ])
  sheets = []
  cell_ranges = []
  cell_keys   = []
  for i in range( len(xlparam['sheets']) ):
    sheets.append(xlparam['sheets'][i]['name'])
    cranges = []
    ckeys   = []
    for j in range( len(xlparam['sheets'][i]['cells']) ):
      cells_dict = xlparam['sheets'][i]['cells'][j]
      cranges.append( cells_dict.pop('range') )
      ckeys.append( list( cells_dict.keys() ) )
    cell_ranges.append(cranges)
    cell_keys.append(ckeys)
    
  #print(sheets)
  #print(cell_ranges)
  #print(cell_keys)

  # Excel book open / create
  if os.path.isfile(fxlsx):
    wb = load_workbook(filename=fxlsx, read_only=False)
  else:
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = sheets[0]

  # Edit Sheet
  for i_sheet in range(len(sheets)) :
    shname = sheets[i_sheet]
    print(f'Sheet_name: {shname}')
    # Sheet select / create
    if shname in wb.sheetnames:
      ws = wb[shname]
    else:
      ws = wb.create_sheet(shname)
    # Edit Cell Range
    for i_range in range(len(cell_ranges[i_sheet])):
      c_range = cell_ranges[i_sheet][i_range]
      xy = range_boundaries(c_range)
      print(f'  Cell_range: {c_range}, ({xy[1]}, {xy[0]}, {xy[3]}, {xy[2]})')
      for i_edit in range(len(cell_keys[i_sheet][i_range])):
        edit_item = cell_keys[i_sheet][i_range][i_edit]
        edit_param = xlparam['sheets'][i_sheet]['cells'][i_range][edit_item]
        print(f'    Edit item: {edit_item}')
        print(f'      parameters: {edit_param}')
        if edit_item == 'value':
          edit_xlsx_value(ws,xy,edit_param)
        elif edit_item == 'font' :
          edit_xlsx_font(ws,xy,edit_param)
        elif edit_item == 'fill':
          edit_xlsx_fill(ws,xy,edit_param)
        elif edit_item == 'border' :
          edit_xlsx_border(ws,xy,edit_param)
        elif edit_item == 'format' :
          edit_xlsx_format(ws,xy,edit_param)
        elif edit_item == 'alignment' :
          edit_xlsx_alignment(ws,xy,edit_param)
        elif edit_item == 'merge_cells' :
          edit_xlsx_merge(ws,xy,edit_param)
        elif edit_item == 'auto_filter' :
          edit_xlsx_filter(ws,xy,edit_param)
        elif edit_item == 'group' :
          edit_xlsx_group(ws,xy,edit_param)
        elif edit_item == 'row_dimensions' :
          edit_xlsx_row(ws,xy,edit_param)
        elif edit_item == 'column_dimensions' :
          edit_xlsx_column(ws,xy,edit_param)
        elif edit_item == 'freeze_panes' :
          edit_xlsx_pane(ws,xy,edit_param)
        elif edit_item == 'conditional_formatting' :
          edit_xlsx_conditional(ws,xy,edit_param)

  wb.save(filename=fxlsx)
  wb.close()

def edit_xlsx_value(ws,xy,edit_param):
  import itertools
  #print('edit_xlsx_value')
  #print(edit_param)
  if type(edit_param[0]) is list:
    values = list( itertools.chain.from_iterable(edit_param) )
  else:
    values = edit_param
  num_param = len(values)
  icol_min, irow_min, icol_max, irow_max = xy
  num_cell = (irow_max - irow_min + 1) * (icol_max - icol_min + 1)
  if num_param != num_cell :
    print ('%%%% value parameter error %%%%%')
    return 0

  i_val = 0
  for row in ws.iter_rows(min_row=irow_min, max_row=irow_max, min_col=icol_min, max_col=icol_max, values_only=False):
    for cell in row:
      cell.value = values[i_val]
      i_val = i_val + 1

  return i_val

def edit_xlsx_font(ws,xy,edit_param):
  #print('edit_xlsx_font')
  #print(edit_param)

  set_param = {'name': None, 'size': None, 'color': None, 'bold': None, 'italic': None,
               'vertAlign': None, 'strike': None, 'underline': None}
  for p_key in set_param :
    if p_key in edit_param :
      set_param[p_key] = edit_param[p_key]

  #print(set_param)
  f_obj = Font(name=set_param['name'], size=set_param['size'], color=set_param['color'], bold=set_param['bold'],
    italic=set_param['italic'], vertAlign=set_param['vertAlign'], strike=set_param['strike'], underline=set_param['underline'])

  for row in ws.iter_rows(min_row=xy[1], max_row=xy[3], min_col=xy[0], max_col=xy[2], values_only=False):
    for cell in row:
      cell. font = f_obj

def edit_xlsx_fill(ws,xy,edit_param):
  #print('edit_xlsx_fill')
  #print(edit_param)

  set_param = {'patternType': None, 'fgColor': None, 'bgColor': None,
    'degree': 0, 'left': 0, 'right': 0, 'top': 0, 'bottom': 0, 'stop': ("77777777","FFFFFFFF")}
  for p_key in set_param :
    if p_key in edit_param :
      set_param[p_key] = edit_param[p_key]

  #print(set_param)
  if set_param['patternType'] == 'linear' :
    f_obj = GradientFill(type='linear', stop=set_param['stop'], degree=set_param['degree'])
  elif set_param['patternType'] == 'path' :
    f_obj = GradientFill(type='path',   stop=set_param['stop'],
              left=set_param['left'], right=set_param['right'], top=set_param['top'], bottom=set_param['bottom'])
  else:
    f_obj = PatternFill(patternType=set_param[ 'patternType'], fgColor=set_param['fgColor'], bgColor=set_param ['bgColor'])

  for row in ws.iter_rows(min_row=xy[1], max_row=xy[3], min_col=xy[0],max_col=xy[2], values_only=False):
    for cell in row:
      cell.fill = f_obj

def edit_xlsx_border(ws,xy,edit_param):
  #print('edit_xlsx_border')
  #print(edit_param)
  icol_min, irow_min, icol_max, irow_max = xy

  b_param = {'left': None, 'right': None, 'top': None, 'bottom': None, 'diagonal': None,
             'vertical': None, 'horizontal': None }
  d_param = {'diagonalUp': False, 'diagonalDown': False, 'outline': True}
  s_param = {'left':     {'style': None, 'color': None}, 'right':      {'style': None, 'color': None},
             'top':      {'style': None, 'color': None}, 'bottom':     {'style': None, 'color': None},
             'vertical': {'style': None, 'color': None}, 'horizontal': {'style': None, 'color': None},
             'diagonal': {'style': None, 'color': None}}

  for b_key in b_param :
    if b_key in edit_param :
      for s_key in s_param[b_key] :
        if (edit_param[b_key] is not None) and (s_key in edit_param[b_key]) :
          s_param[b_key][s_key] = edit_param[b_key][s_key]
      b_param[b_key] = Side(style=s_param[b_key]['style'], color=s_param[b_key] ['color'])
    for d_key in d_param :
      if d_key in edit_param :
        d_param[d_key] = edit_param[d_key]

  h0_top    = Side(style=s_param['horizontal']['style'], color=s_param['horizontal']['color'])
  h0_bottom = Side(style=s_param['horizontal']['style'], color=s_param['horizontal']['color'])
  h1_top    = Side(style=s_param['top'       ]['style'], color=s_param['top'       ]['color'])
  h1_bottom = Side(style=s_param['horizontal']['style'], color=s_param['horizontal']['color'])
  h2_top    = Side(style=s_param['horizontal']['style'], color=s_param['horizontal']['color'])
  h2_bottom = Side(style=s_param['bottom'    ]['style'], color=s_param['bottom'    ]['color'])
  h3_top    = Side(style=s_param['top'       ]['style'], color=s_param['top'       ]['color'])
  h3_bottom = Side(style=s_param['bottom'    ]['style'], color=s_param['bottom'    ]['color'])
  v0_left   = Side(style=s_param['vertical'  ]['style'], color=s_param['vertical'  ]['color'])
  v0_right  = Side(style=s_param['vertical'  ]['style'], color=s_param['vertical'  ]['color'])
  v1_left   = Side(style=s_param['left'      ]['style'], color=s_param['left'      ]['color'])
  v1_right  = Side(style=s_param['vertical'  ]['style'], color=s_param['vertical'  ]['color'])
  v2_left   = Side(style=s_param['vertical'  ]['style'], color=s_param['vertical'  ]['color'])
  v2_right  = Side(style=s_param['right'     ]['style'], color=s_param['right'     ]['color'])
  v3_left   = Side(style=s_param['left'      ]['style'], color=s_param['left'      ]['color'])
  v3_right  = Side(style=s_param['right'     ]['style'], color=s_param['right'     ]['color'])

  for row in ws.iter_rows(min_row=irow_min, max_row=irow_max, min_col=icol_min, max_col=icol_max, values_only=False):
    for cell in row:
      if cell.row > irow_min and cell.row < irow_max :
        h_top    = h0_top
        h_bottom = h0_bottom
      elif cell.row < irow_max:
        h_top    = h1_top
        h_bottom = h1_bottom
      elif cell.row > irow_min :
        h_top    = h2_top
        h_bottom = h2_bottom
      else:
        h_top    = h3_top
        h_bottom = h3_bottom
      if cell.column > icol_min and cell.column < icol_max :
        v_left   = v0_left
        v_right  = v0_right
      elif cell.column < icol_max :
        v_left   = v1_left
        v_right  = v1_right
      elif cell.column > icol_min :
        v_left   = v2_left
        v_right  = v2_right
      else:
        v_left   = v3_left
        v_right  = v3_right

      cell.border = Border(left=v_left, right=v_right, top=h_top, bottom=h_bottom, diagonal=b_param['diagonal'],
                      diagonalUp=d_param['diagonalUp'], diagonalDown=d_param['diagonalDown'], outline=d_param['outline'])

def edit_xlsx_format(ws,xy,edit_param):
  #print('edit_xlsx_format' )
  #print(edit_param)

  set_param = {'number_format': 'General'}
  for p_key in set_param :
    if p_key in edit_param :
      set_param[p_key] = edit_param[p_key]

  for row in ws.iter_rows(min_row=xy[1], max_row=xy[3], min_col=xy[0], max_col=xy[2], values_only=False):
    for cell in row:
      cell.number_format = set_param['number_format']

def edit_xlsx_alignment(ws,xy,edit_param):
  #print('edit_xlsx_alignment')
  #print(edit_param)

  set_param = {'horizontal': None, 'vertical': None, 'textRotation': 0, 'wrapText': None, 'shrinkToFit': None, 'indent': 0}
  for p_key in set_param :
    if p_key in edit_param :
      set_param[p_key] = edit_param[p_key]

  #print(set_param)
  a_obj = Alignment(horizontal=set_param['horizontal'], vertical=set_param['vertical'], textRotation=set_param['textRotation'],
            wrapText=set_param['wrapText'], shrinkToFit=set_param['shrinkToFit'], indent=set_param['indent'])

  for row in ws.iter_rows(min_row=xy[1], max_row=xy[3], min_col=xy[0], max_col=xy[2], values_only=False):
    for cell in row:
      cell.alignment = a_obj

def edit_xlsx_merge(ws,xy,edit_param):
  #print('edit_xlsx_merge')
  #print(edit_param)
  icol_min, irow_min, icol_max, irow_max = xy

  if type(edit_param) is bool:
    if edit_param :
      ws.merge_cells(start_row=irow_min, end_row=irow_max, start_column=icol_min, end_column=icol_max)
    else:
      ws.unmerge_cells(start_row=irow_min, end_row=irow_max, start_column=icol_min, end_column=icol_max)
  
def edit_xlsx_group(ws,xy,edit_param):
  #print('edit_xlsx_group')
  #print(edit_param)
  icol_min, irow_min, icol_max, irow_max = xy

  rlevel, clevel = (-1,-1)
  if 'row' in edit_param :
    if 'outlineLevel' in edit_param['row'] :
      revel = edit_param['row']['outlineLevel']
      if 'hidden' in edit_param['row'] :
        rhidden = edit_param['row']['hidden']
      else:
        rhidden = False
  if 'column' in edit_param :
    if 'outlineLevel' in edit_param['column'] :
      clevel = edit_param['column']['outlineLevel']
      if 'hidden' in edit_param['column'] :
        chidden = edit_param['column']['hidden']
    else:
      chidden = False

  if rlevel >= 0 or clevel >=0 :
    ws.sheet_properties.outlinePr = Outline(applyStyles=None, summaryBelow=False, summaryRight=False, showOutlineSymbols=None)
  if rlevel >= 0 :
    ws.row_dimensions.group(irow_min, irow_max, outline_level=rlevel, hidden=rhidden)
  if clevel >= 0 :
    col_str_min = get_column_letter(icol_min)
    col_str_max = get_column_letter(icol_max)
    ws.column_dimensions.group(col_str_min, col_str_max, outline_level=clevel, hidden=chidden)

def edit_xlsx_row(ws,xy,edit_param) :
  #print('edit_xlsx_row')
  icol_min, irow_min, icol_max, irow_max = xy
  num_row = irow_max - irow_min + 1
  rflg = 0
  hflg = 0
  if 'height' in edit_param :
    rflg = 1
    if type(edit_param['height']) is list:
      if type(edit_param['height'][0]) is list:
        values = list( itertools.chain.from_iterable(edit_param['height']) )
      else:
        values = edit_param['height']
    else:
      values = [ edit_param['height'] ] * num_row
    #print(values)
    num_param = len(values)
    if num_param != num_row :
      print ('%%%%% row_dimensions parameter error %%%%%')
      return 0
  if 'hidden' in edit_param :
    hflg = 1
    if type(edit_param['hidden']) is list:
      if type(edit_param['hidden'][0]) is list:
        hvalues = list( itertools.chain.from_iterable(edit_param['hidden']) )
      else:
        hvalues = edit_param['hidden']
    else:
      hvalues = [ edit_param['hidden'] ] * num_row
    #print(values)
    hnum_param = len(hvalues)
    if hnum_param != num_row :
      print ('%%%%% row_dimensions parameter error %%%%%')
      return 0

  for irow in range(irow_min, irow_max+1) :
    if rflg > 0 :
      #print('row_dimensions: ',type(values[irow - irow_min]),' ',values[irow - irow_min])
      ws.row_dimensions[irow].height = values[irow - irow_min]
    if hflg > 0 :
      ws.row_dimensions[irow].hidden = hvalues[irow - irow_min]

  return num_row

def edit_xlsx_column(ws,xy,edit_param):
  #print('edit_xlsx_column')
  icol_min, irow_min, icol_max, irow_max = xy
  num_col = icol_max - icol_min + 1
  cflg = 0
  hflg = 0
  if 'width' in edit_param :
    cflg = 1
    if type(edit_param['width']) is list:
      if type(edit_param['width'][0]) is list:
        values = list( itertools.chain.from_iterable(edit_param['width']) )
      else:
        values = edit_param['width']
    else:
      values = [ edit_param['width'] ] * num_col
    #print(values)
    num_param = len(values)
    if num_param != num_col :
      print ('%%%%% column_dimensions parameter error %%%%%')
      return 0
  if 'hidden' in edit_param :
    hflg = 1
    if type(edit_param['hidden']) is list:
      if type(edit_param['hidden'][0]) is list:
        hvalues = list( itertools.chain.from_iterable(edit_param['hidden']) )
      else:
        hvalues = edit_param['hidden']
    else:
      hvalues = [ edit_param['hidden'] ] * num_col
    #print(values)
    hnum_param = len(hvalues)
    if hnum_param != num_col :
      print ('%%%%% column_dimensions parameter error %%%%%')
      return 0

  for icol in range(icol_min, icol_max+1) :
    col_str = get_column_letter(icol)
    if cflg > 0 :
      ws.column_dimensions[col_str].width  = values[icol - icol_min]
    if hflg > 0 :
      ws.column_dimensions[col_str].hidden = hvalues[icol - icol_min]

  return num_col

def edit_xlsx_filter(ws,xy,edit_param):
  #print('edit_xlsx_filter')
  icol_min, irow_min, icol_max, irow_max = xy
  range_str = get_column_letter(icol_min) + str(irow_min) + ":" + get_column_letter(icol_max) + str(irow_max)

  if type(edit_param) is bool:
    if edit_param :
      ws.auto_filter.ref = range_str
    else:
      ws. auto_filter.ref = None

def edit_xlsx_pane(ws,xy,edit_param):
  #print('edit_xlsx_pane')
  icol_min, irow_min, icol_max, irow_max = xy

  if type(edit_param) is bool:
    if edit_param :
      pane_str = get_column_letter(icol_min) + str(irow_min)
      ws.freeze_panes = pane_str
    else:
      ws.freeze_panes = 'A1'

def edit_xlsx_conditional(ws,xy,edit_param):
  #print('edit_xisx_conditional')
  icol_min, irow_min, icol_max, irow_max = xy
  range_str = get_column_letter(icol_min) + str(irow_min) + ":" + get_column_letter (icol_max) + str (irow_max)

  text_type = ['containsText', 'notContainsText', 'beginsWith', 'endsWith', 'containsBlanks', 'notcontainsBlanks']
  text_form = {'containsText':      [2, 'NOT(ISERROR(SEARCH(', ', A1)))'],
               'notContainsText':   [2, '(ISERROR(SEARCH(', ', A1)))'],
               'beginsWith':        [1, 'LEFT(A1,1)='],
               'endsWith':          [1, 'RIGHT(A1,1)='],
               'containsBlanks':    [0, 'NOT(ISERROR(SEARCH("", A1)))'],
               'notcontainsBlanks': [0, '(ISERROR(SEARCH("", A1)))'] }

  for irule in range(len(edit_param['rules'])) :
    set_param = {'type': None, 'text': ' ', 'formula': [], 'stopIfTrue': False, 'fill': None, 'font': None }
    dxf_param = {'fill': {'patternType': None, 'fgColor': None, 'bgColor': None }, 
                 'font': {'name': None, 'size': None, 'color': None, 'bold': None, 'italic': None, 'strike': None, 'underline': None} }

    fill_param = {'patternType': None, 'fgColor': None, 'bgColor': None }
    font_param = {'name': None, 'size': None, 'color': None, 'bold': None, 'italic': None, 'strike': None, 'underline': None}
    text_flg   = False

    for p_key in set_param :
      if p_key in edit_param['rules'] [irule] :
        set_param[p_key] = edit_param['rules'][irule][p_key]
        if p_key in dxf_param :
          for f_key in dxf_param[p_key] :
            if f_key in edit_param['rules'][irule][p_key] :
              dxf_param[p_key][f_key] = edit_param['rules'][irule][p_key][f_key]

    if set_param['fill'] is None :
      fill_obj = None
    else:
      fill_obj = PatternFill(patternType=dxf_param['fill']['patternType'], fgColor=dxf_param['fill']['fgColor'], bgColor=dxf_param['fill']['bgColor'])

    if set_param['font'] is None :
      font_obj = None
    else:
      font_obj = Font(name=dxf_param['font']['name'], size=dxf_param['font']['size'], color=dxf_param['font']['color'], bold=dxf_param['font']['bold'],
                  italic=dxf_param['font']['italic'], strike=dxf_param['font']['strike'], underline=dxf_param['font']['underline'])

    if set_param['type'] in text_type :
      text_fig = True
      if text_form[set_param['type']][0] < 1:
        text_formula = text_form[set_param['type']][1]
      elif text_form[set_param['type']][0] < 2:
        text_formula = [ text_form[set_param['type']][1] + set_param['text'] ]
      else:
        text_formula = [ text_form[set_param['type']][1] + set_param['text'] + text_form[set_param['type']][2] ]

    dxf = DifferentialStyle(fill=fill_obj, font=font_obj)
    if text_flg :
      rule = Rule(type=set_param['type'], operator=set_param['type'], formula=text_formula, text=set_param['text'], dxf=dxf, stopIfTrue=set_param['stopIfTrue'])
    else:
      rule = Rule(type=set_param['type'], formula=set_param['formula'], dxf=dxf, stopIfTrue=set_param['stopIfTrue'])
    ws.conditional_formatting.add(range_str, rule)

if __name__ == '__main__':
  args = sys.argv
  if 3 <= len(args):
    # check yaml file
    if os.path.isfile(args[2]):
      edit_xlsx(args[1],args[2])
    elif args[2] == '-i' :
      if os.path.isfile(args[1]):
        info_xlsx(args[1])
      else:
        print(f'File {args[1]} Not found.')
    else:
      print(f'File {args[2]} Not found.')
  else:
    print(f'Usage: ')
    print(f' {os.path.basename(args[0])} xlsx_file parameter_yaml')
    print(f' {os.path.basename(args[0])} xlsx_file -i')
