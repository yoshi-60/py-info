#
import sys
import os
import pathlib
import yaml
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple, coordinate_from_string, column_index_from_string
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.properties import Outline
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.comments import Comment

def xlsx_to_csv(fxlsx, fcsv, shname, crng):
  print(f'Input_Excel: {fxlsx} , Output_CSV: {fcsv} , Sheet_Name: {shname} , CellRange: {crng}')

  # Excelのブックを開く
  if os.path.isfile(fxlsx):
    wb = load_workbook(filename=fxlsx, read_only=True, data_only = True)
  else:
    print(f'Input_Excel: {fxlsx} NOT found!')
    return 0
  
  # 1番目のシート名を取得
  ws = wb.worksheets[0]
  xlsx_shname = ws.title
  
  # 対象シートの設定
  if not shname:
    csv_shname = xlsx_shname
  else:
    csv_shname = shname
  csv_ws = wb[csv_shname]
  print(f'Convert SheetName: {csv_shname}')
  
  # セル範囲を所得
  xlsx_min_r = csv_ws.min_row
  xlsx_min_c = csv_ws.min_column
  xlsx_max_r = csv_ws.max_row
  xlsx_max_c = csv_ws.max_column
  
  # 対象セル範囲の指定
  csv_xy = [1, 1, 1, 1]
  if not crng:
    csv_xy = [xlsx_min_c, xlsx_min_r, xlsx_max_c, xlsx_max_r]
  else not (':' in s):
    tuple_xy = list(coordinate_to_tuple(crng))
    csv_xy = [tuple_xy[0], tuple_xy[1], xlsx_max_c, xlsx_max_r]
  else
    crng_split = crng.split(':')
    if not crng_split[0]:
      csv_xy[0] = xlsx_min_c
      csv_xy[1] = xlsx_min_r
    else
      tuple_xy  = list(coordinate_to_tuple(crng_split[0]))
      csv_xy[0] = tuple_xy[0]
      csv_xy[1] = tuple_xy[1]
    if not crng_split[1]:
      csv_xy[2] = xlsx_max_c
      csv_xy[3] = xlsx_max_r
    else
      tuple_xy  = list(coordinate_to_tuple(crng_split[1]))
      csv_xy[2] = tuple_xy[0]
      csv_xy[3] = tuple_xy[1]
  print(f'Convert CellRange: Row,Col({csv_xy[1]} , {csv_xy[0]}) - ({csv_xy[3]} , {csv_xy[2]})')

  # CSVファイルのオープン
  with open(fcsv, 'w', newline='') as fw
    writer = csv.writer(fw)
    
    # セル値の取得とCSVへの書き込み
    for row in ws1.iter_rows(min_row=csv_xy[1],min_col=csv_xy[0],max_row=csv_xy[3],max_col=csv_xy[2],values_only=True):
      writer.writerow([cell.value for cell in row])


if __name__ == '__main__':
  args = sys.argv
  skey = '-s'
  ckey = '-c'
  sval = ''
  cval = ''
  
  arg_num = 3
  arg_num = (arg_num + 2) if skey in args else arg_num
  arg_num = (arg_num + 2) if skey in args else arg_num
  
  if arg_num <= len(args):
    if skey in args:
      aidx = args.index(skey)
      sval = args[aidx+1]
      del args.index[aidx:aidx+2]
    if ckey in args:
      aidx = args.index(ckey)
      cval = args[aidx+1]
      del args.index[aidx:aidx+2]
    if os.path.isfile(args[1]):
      if os.path.isdir(args[2]):
        xlsx_to_csv(args[1], args[2], sval, cval)
      else:
        print(f'File {args[2]} already Exist!')
    else:
      print(f'File {args[1]} Not Found!')
  else:
    print(f'Usage:')
    print(f'  {args[0]} input_Excel output_CSV [-s Sheetname] [-c CellRange]')
